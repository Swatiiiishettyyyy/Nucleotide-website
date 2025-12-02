from functools import lru_cache
from pathlib import Path
from typing import Optional, Set
import logging

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore
    logger.error("openpyxl is required to read Locations.xlsx. Please install it.")


LOCATIONS_FILE = Path(__file__).resolve().parents[1] / "Locations.xlsx"


def _clear_location_cache():
    """Clear the cache for serviceable locations (useful for testing/reloading)"""
    _load_serviceable_locations.cache_clear()


@lru_cache(maxsize=1)
def _load_serviceable_locations() -> Set[str]:
    """Load allowed locations (cities/localities) from Locations.xlsx."""
    allowed: Set[str] = set()
    if load_workbook is None:
        return allowed

    if not LOCATIONS_FILE.exists():
        logger.error("Locations.xlsx file not found at %s", LOCATIONS_FILE)
        return allowed

    try:
        workbook = load_workbook(LOCATIONS_FILE, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover
        logger.error("Failed to read Locations.xlsx: %s", exc)
        return allowed

    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        logger.error("Locations.xlsx is empty.")
        return allowed

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]
    try:
        location_idx = headers.index("location")
    except ValueError:
        logger.error("'Location' column not found in Locations.xlsx headers.")
        return allowed

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        value = row[location_idx]
        if value:
            normalized = str(value).strip().lower()
            if normalized:
                allowed.add(normalized)

    logger.info("Loaded %d serviceable locations from Locations.xlsx", len(allowed))
    return allowed


def is_serviceable_location(city: Optional[str], locality: Optional[str] = None) -> bool:
    """
    Return True if city name is present in the Locations.xlsx list.
    Only city name is validated from the excel sheet.
    """
    allowed = _load_serviceable_locations()
    if not allowed:
        # If file missing or empty, be conservative and reject.
        logger.warning("Locations.xlsx is missing or empty. Rejecting all locations.")
        return False

    def normalize(value: Optional[str]) -> Optional[str]:
        if not value:
            return None
        if not isinstance(value, str):
            return None
        normalized = value.strip().lower()
        return normalized if normalized else None

    normalized_city = normalize(city)

    # Log for debugging
    logger.info(f"Validating city name - City: '{city}' (normalized: '{normalized_city}')")
    logger.info(f"Total serviceable locations loaded: {len(allowed)}")

    # Check city name in excel sheet
    if normalized_city:
        if normalized_city in allowed:
            logger.info(f"City '{city}' (normalized: '{normalized_city}') found in serviceable locations.")
            return True
        else:
            logger.warning(f"City '{city}' (normalized: '{normalized_city}') NOT found in serviceable locations.")
            # Check if there's a similar city name (for debugging)
            similar = [loc for loc in allowed if normalized_city in loc or loc in normalized_city]
            if similar:
                logger.info(f"Similar city names found in Excel: {similar[:5]}")
    
    # City not found in serviceable locations
    logger.warning(f"Location validation failed - City: '{city}' (normalized: '{normalized_city}')")
    return False


