"""
One-off script to populate serviceable_locations from Locations.xlsx.

Run from project root:
    python seed_serviceable_locations_from_xlsx.py [path/to/Locations.xlsx]

If no path is given, looks for Locations.xlsx in the project root.
Expects the Excel file to have a header row with a column named "location" (case-insensitive).
"""
import sys
from pathlib import Path

# Ensure project root is on path
PROJECT_ROOT = Path(__file__).resolve().parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

def main() -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    from database import SessionLocal
    from Address_module.Address_model import ServiceableLocation

    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else PROJECT_ROOT / "Locations.xlsx"
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        print("Excel file has no header row.")
        sys.exit(1)

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]
    try:
        location_idx = headers.index("location")
    except ValueError:
        print("'location' column not found. Headers:", headers)
        sys.exit(1)

    allowed: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        value = row[location_idx]
        if value:
            normalized = str(value).strip().lower()
            if normalized:
                allowed.add(normalized)

    workbook.close()
    if not allowed:
        print("No location values found in the file.")
        sys.exit(0)

    db = SessionLocal()
    try:
        existing = {r.location for r in db.query(ServiceableLocation.location).all()}
        to_add = [loc for loc in sorted(allowed) if loc not in existing]
        for loc in to_add:
            db.add(ServiceableLocation(location=loc))
        db.commit()
        print(f"Inserted {len(to_add)} locations. Skipped {len(allowed) - len(to_add)} already present. Total rows in table: {len(existing) + len(to_add)}.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
